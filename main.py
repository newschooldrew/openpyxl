import openpyxl
from openpyxl.utils.cell import get_column_letter

# wb = openpyxl.Workbook()
# wb.save('first_wb.xlsx')
#
# wb = openpyxl.load_workbook('first_wb.xlsx')
#
# for sheet in wb:
#     print(sheet.title)

############################################

# wb = Workbook()
# ws = wb.create_sheet('A Sheet',0)
# ws2 = wb.create_sheet('B Sheet')
#
# wb.remove(wb['Sheet'])
# del wb['A Sheet']
#
# ws2.title = "new title"
# wb.save('new_wb.xlsx')
# for sheet in wb:
#     print(sheet.title)

################################################

# wb = load_workbook('new_wb.xlsx')
# for sheet in wb:
#     print(sheet.title)
#
# source = wb['new title']
# new_sheet = wb.copy_worksheet(source)
# new_sheet.title = 'copied title'
# for sheet in wb:
#     print(sheet.title)
# wb.save('copied_sheets.xlsx')

####################################

# wb = openpyxl.Workbook()
# ws = wb.worksheets[0]
# ws['A1'].value = 56
# ws['C3'].value = 'abc'
# ws.cell(row=1,column=4).value = 111
# ws.cell(row=5,column=4).value = 3333
# wb.save('fill_in_cells.xlsx')

####################################
# LOOPING
####################################
wb = openpyxl.Workbook()
ws = wb.worksheets[0]

for row in range(1, 4):
    for col in range(1, 4):
        cell = ws.cell(row=row, column=col)
        cell.value = 1


####################################
# LOOPING
####################################
# wb = openpyxl.Workbook()
# ws = wb.worksheets[0]
#
# for row in ws.iter_rows(min_row=1,min_col=1,max_col=4,max_row=3):
#     print(row)
#     # for cell in row:
#     #     print(cell.coordinate, end=" ")
#
# for col in ws.iter_cols(max_col=4,max_row=3):
#     print(col)
#     for cell in col:
#         print(cell.coordinate,end="")

####################################
# INSERTING FORMULAS
####################################

wb = openpyxl.Workbook()
ws = wb.worksheets[0]

def print_rows(ws):
    row_string = ""
    for row in ws.iter_rows(min_row=1,max_col=ws.max_column,max_row=ws.max_row):
        for cell in row:
            row_string += "{:<30}".format(str(cell.value) + " ")
        row_string += "\n"
    print(row_string)

def set_values(ws):
    ws.delete_cols(1,100)
    counter = 1
    for row in ws.iter_rows(min_row=1,max_col=10,max_row=10):
        for cell in row:
            cell.value = counter
            counter+= 1

for i in range(1,11):
    ws.cell(row=i,column=1).value = i * i
    ws.cell(row=i,column=2).value = i/2

first_cell = ws.cell(row=1,column=1)
last_cell = ws.cell(row=10,column=1)

first_cell = ws.cell(row=1,column=2)
last_cell = ws.cell(row=10,column=2)

ws.cell(row=11,column=1).value = \
    "=SUM(" + \
    str(first_cell.coordinate) + \
    ":" + \
    str(last_cell.coordinate) + \
    ")"

ws.cell(row=11,column=2).value = \
    "=SUM(" + \
    str(first_cell.coordinate) + \
    ":" + \
    str(last_cell.coordinate) + \
    ")"

# set_values(ws)
print_rows(ws)

wb.save('looping.xlsx')
